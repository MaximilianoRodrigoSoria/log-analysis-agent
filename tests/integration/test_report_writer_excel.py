import sys
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from src.adapters.report_writer_excel import ExcelReportWriter


def test_excel_writer_creates_formatted_report():
    analysis = {
        "summary": {
            "total_events": 2,
            "total_errors": 1,
            "total_warnings": 1
        },
        "error_groups": [
            {
                "exception": "NullPointerException",
                "count": 1,
                "logger": "com.example",
                "first_ts": "2026-02-13 10:00:00",
                "last_ts": "2026-02-13 10:00:00"
            }
        ],
        "warnings": [
            {
                "ts": "2026-02-13 10:01:00",
                "logger": "com.example",
                "message": "warn"
            }
        ]
    }

    with TemporaryDirectory() as temp_dir:
        writer = ExcelReportWriter(Path(temp_dir))
        path = writer.write_report("run1", report_content="", analysis=analysis)

        workbook = load_workbook(path)
        assert set(workbook.sheetnames) == {"summary", "error_groups", "warnings"}

        summary_sheet = workbook["summary"]
        header_cell = summary_sheet.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert summary_sheet.column_dimensions["A"].width >= 12
