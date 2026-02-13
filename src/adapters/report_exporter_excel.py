"""
Exporter de reportes en formato Excel (.xlsx).
Genera hojas de cálculo con tablas estructuradas y formato profesional.
"""

import logging
from pathlib import Path
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

from ..ports.report_exporter_port import ReportExporterPort


logger = logging.getLogger(__name__)


class ExcelExporter(ReportExporterPort):
    """Exporta reportes en formato Excel con formato profesional"""
    
    def export(
        self,
        output_dir: str,
        output_filename: str,
        report_content: str,
        analysis: Optional[Dict] = None
    ) -> str:
        """
        Exporta el reporte en formato Excel (.xlsx).
        Crea hojas con: Resumen, Errores, Warnings.
        
        Args:
            output_dir: Directorio de salida
            output_filename: Nombre del archivo (sin extensión)
            report_content: Contenido del reporte (no usado, Excel usa analysis)
            analysis: Análisis estructurado (requerido)
        
        Returns:
            Path absoluto del archivo generado
        
        Raises:
            IOError: Si hay error de escritura
            ValueError: Si no hay analysis disponible
        """
        if analysis is None:
            raise ValueError("Análisis estructurado requerido para exportar a Excel")
        
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        filename = f"{output_filename}.xlsx"
        file_path = output_path / filename
        
        logger.debug(f"Exportando reporte Excel a {file_path}")
        
        try:
            workbook = Workbook()
            
            # Remover hoja por defecto
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Crear hojas
            self._create_summary_sheet(
                workbook, 
                analysis, 
                header_font, 
                header_fill, 
                thin_border,
                center_alignment
            )
            
            self._create_errors_sheet(
                workbook, 
                analysis, 
                header_font, 
                header_fill, 
                thin_border
            )
            
            self._create_warnings_sheet(
                workbook, 
                analysis, 
                header_font, 
                header_fill, 
                thin_border
            )
            
            # Guardar archivo
            workbook.save(file_path)
            
            logger.info(f"Reporte Excel exportado: {file_path}")
            return str(file_path.absolute())
        
        except Exception as e:
            logger.error(f"Error al exportar reporte Excel: {e}")
            raise IOError(f"Error al escribir archivo Excel: {e}") from e
    
    def _create_summary_sheet(
        self,
        workbook: Workbook,
        analysis: Dict,
        header_font: Font,
        header_fill: PatternFill,
        border: Border,
        alignment: Alignment
    ):
        """Crea hoja de resumen"""
        sheet = workbook.create_sheet("Resumen", 0)
        
        # Encabezados
        sheet['A1'] = "Métrica"
        sheet['B1'] = "Valor"
        
        sheet['A1'].font = header_font
        sheet['A1'].fill = header_fill
        sheet['A1'].border = border
        sheet['A1'].alignment = alignment
        
        sheet['B1'].font = header_font
        sheet['B1'].fill = header_fill
        sheet['B1'].border = border
        sheet['B1'].alignment = alignment
        
        # Datos del resumen
        if "summary" in analysis:
            row = 2
            for key, value in analysis["summary"].items():
                label = key.replace('_', ' ').title()
                sheet[f'A{row}'] = label
                sheet[f'B{row}'] = value
                sheet[f'A{row}'].border = border
                sheet[f'B{row}'].border = border
                sheet[f'B{row}'].alignment = alignment
                row += 1
        
        # Autoajustar ancho de columnas
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
    
    def _create_errors_sheet(
        self,
        workbook: Workbook,
        analysis: Dict,
        header_font: Font,
        header_fill: PatternFill,
        border: Border
    ):
        """Crea hoja de errores detallados"""
        sheet = workbook.create_sheet("Errores")
        
        # Encabezados
        headers = ["Tipo de Error", "Ocurrencias", "Componente", "Ubicación", "Primera Vez", "Última Vez"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Datos de errores
        if "error_groups" in analysis:
            row = 2
            for error_group in analysis["error_groups"]:
                exception = error_group.get("exception", "Unknown")
                count = error_group.get("count", 0)
                component = error_group.get("logger", "Unknown")
                
                # Ubicación del error
                location = ""
                if "top_frame" in error_group and error_group["top_frame"]:
                    frame = error_group["top_frame"]
                    location = f"{frame.get('file', 'unknown')}:{frame.get('line', '?')}"
                
                first_ts = error_group.get("first_ts", "N/A")
                last_ts = error_group.get("last_ts", "N/A")
                
                sheet[f'A{row}'] = exception
                sheet[f'B{row}'] = count
                sheet[f'C{row}'] = component
                sheet[f'D{row}'] = location
                sheet[f'E{row}'] = first_ts
                sheet[f'F{row}'] = last_ts
                
                for col in range(1, 7):
                    sheet.cell(row=row, column=col).border = border
                
                row += 1
        
        # Autoajustar columnas
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 35
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
    
    def _create_warnings_sheet(
        self,
        workbook: Workbook,
        analysis: Dict,
        header_font: Font,
        header_fill: PatternFill,
        border: Border
    ):
        """Crea hoja de warnings"""
        sheet = workbook.create_sheet("Advertencias")
        
        # Encabezados
        headers = ["Timestamp", "Componente", "Mensaje"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Datos de warnings
        if "warnings" in analysis:
            row = 2
            for warning in analysis["warnings"][:50]:  # Limitar a 50
                timestamp = warning.get("timestamp", "N/A")
                component = warning.get("logger", "Unknown")
                message = warning.get("message", "")
                
                sheet[f'A{row}'] = timestamp
                sheet[f'B{row}'] = component
                sheet[f'C{row}'] = message
                
                for col in range(1, 4):
                    sheet.cell(row=row, column=col).border = border
                
                row += 1
        
        # Autoajustar columnas
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 60
