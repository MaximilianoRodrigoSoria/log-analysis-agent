"""
Adapter para generar reportes Excel.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

from ..config.constants import Constants
from ..domain.log_analyzer.report_schema import get_report_schema


logger = logging.getLogger(__name__)


class ExcelReportWriter:
    """Genera reportes Excel con formato profesional"""

    def __init__(self, reports_dir: Path):
        self.reports_dir = reports_dir

    def write_report(
        self,
        run_id: str,
        report_content: str,
        analysis: Optional[Dict] = None
    ) -> str:
        """
        Escribe el reporte en formato Excel.

        Args:
            run_id: Identificador de la ejecucion
            report_content: Contenido Markdown (no usado en Excel)
            analysis: Analisis estructurado

        Returns:
            Path absoluto del archivo generado
        """
        if analysis is None:
            raise ValueError("Analisis requerido para generar Excel")

        self.reports_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{run_id}{Constants.REPORT_FILE_EXTENSION_XLSX}"
        filepath = self.reports_dir / filename

        workbook = Workbook()
        workbook.remove(workbook.active)

        schema = get_report_schema()
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        header_font = Font(bold=True)

        self._write_section(
            workbook,
            "summary",
            schema.get("summary", []),
            [analysis.get("summary", {})],
            header_font,
            thin_border
        )
        self._write_section(
            workbook,
            "error_groups",
            schema.get("error_groups", []),
            analysis.get("error_groups", []),
            header_font,
            thin_border
        )
        self._write_section(
            workbook,
            "warnings",
            schema.get("warnings", []),
            analysis.get("warnings", []),
            header_font,
            thin_border
        )

        workbook.save(filepath)
        return str(filepath.resolve())

    def _write_section(
        self,
        workbook: Workbook,
        title: str,
        columns: List[Dict[str, str]],
        rows: List[Dict],
        header_font: Font,
        border: Border
    ) -> None:
        sheet = workbook.create_sheet(title=title)

        for col_index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_index, value=column["label"])
            cell.font = header_font
            cell.border = border

        for row_index, row in enumerate(rows, start=2):
            for col_index, column in enumerate(columns, start=1):
                value = row.get(column["key"], "")
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border

        self._auto_fit_columns(sheet, min_width=12)

    @staticmethod
    def _auto_fit_columns(sheet, min_width: int) -> None:
        for col in sheet.columns:
            max_len = min_width
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_len
